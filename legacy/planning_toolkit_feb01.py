import yaml
import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, NamedStyle, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
import datetime
import os
import logging
import sys
import json
import psutil
import subprocess


## configuration
# path of this python file
file_path = os.path.abspath(__file__)
config_file_name = 'calendar_config_2025_feb01.yaml'
config_file = os.path.join(os.path.dirname(file_path), config_file_name)
    

## globals
# datestamp
ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
ds = datetime.datetime.now().strftime('%Y%m%d')

## logging
# set log file
path = os.path.dirname(file_path)
log_path = os.path.join(path,'logs')
log_file_name = 'calendar_{ts}.log'.format(ts=ts)
log_file = os.path.join(log_path, log_file_name)
# make logs directory if it doesn't exist
if not os.path.exists(log_path):
    os.makedirs(log_path)
    
# configure logging
log = logging.getLogger()
log.setLevel(logging.DEBUG)
logFormatter = logging.Formatter("%(asctime)s [%(threadName)-12.12s] [%(levelname)-5.5s]  %(message)s")
# file output
fileHandler = logging.FileHandler(log_file)
fileHandler.setFormatter(logFormatter)
fileHandler.setLevel(logging.DEBUG)
# console output
consoleHandler = logging.StreamHandler((sys.stdout))
consoleHandler.setFormatter(logFormatter)
consoleHandler.setLevel(logging.INFO)
# add handlers
log.addHandler(consoleHandler)
log.addHandler(fileHandler)

# test logging:
# log.debug('Logging - debug')
# log.info('Logging - info')
# log.warning('Logging - warning')
# log.error('Logging - error')

## global holders
config = {}
wb = None
ws = None
xlsx_file = None
event_dates = {}

def load_config(config_file):
    """
    Load configuration from a YAML file.

    Args:
        config_file (str): The path to the configuration file.

    Globals:
        log (logging.Logger): The logger object.

    Returns:
        dict: The loaded configuration as a dictionary.
    """
    log.info(f'Loading configuration from {config_file}')
    with open(config_file, 'r') as file:
        return yaml.safe_load(file)


def create_calendar(xlsx_file, worksheet_name, overwrite=True):
    """
    Creates a calendar by creating a new workbook, worksheet, and saving it as an Excel file.

    Globals:
        log (logging.Logger): The logger object.

    Returns:
        Workbook: The created workbook object.

    Raises:
        FileExistsError: If the Excel file already exists and overwrite is set to False.
    """
    log.info('Creating calendar')
    
    # create workbook
    log.info(f'Creating calendar {xlsx_file}')
    if os.path.exists(xlsx_file):
        if not overwrite: 
            log.error(f'File {xlsx_file} already exists. Exiting.')
            raise FileExistsError(f'File {xlsx_file} already exists. Overrite set to False. Exiting.') 
        log.warning(f'File {xlsx_file} already exists. Overwriting it.')
    wb = openpyxl.Workbook()

    # configure worksheet
    ws = wb.active
    ws.title = ws_name

    # initialise the first 10 rows and 5 columns
    for x in range(1,201):
       for y in range(1,6):
           ws.cell(row=x, column=y)

    wb.save(xlsx_file)
    return wb

def save(wb=wb, xlsx_file=xlsx_file):
    """
    Save the workbook as an Excel file.

    Args:
        wb (Workbook): The workbook object.
        xlsx_file (str): The path to the Excel file.

    Globals:
        log (logging.Logger): The logger object.
        config (dict): The configuration dictionary.
    
    Raises:
        FileNotFoundError: If the Excel file does not exist.    
    """
    log.info('Saving workbook')
    # if the workbook is not provided, use the global one
    if not xlsx_file:
        xlsx_file = config['calendar']['xlsx_name']
    if os.path.exists(xlsx_file):
        log.info(f'Saving workbook as {xlsx_file}')
        wb.save(xlsx_file)
    else:
        log.error(f'File {xlsx_file} does not exist. Exiting.')
        raise FileNotFoundError(f'File {xlsx_file} does not exist. Exiting.')

def is_workbook_open(file_path):
    """
    Check if a specific Excel workbook is currently open.

    Args:
        file_path (str): The full path to the Excel workbook file.

    Returns:
        bool: True if the workbook is open, False otherwise.

    Raises:
        psutil.NoSuchProcess: If the process no longer exists.
        psutil.AccessDenied: If the process cannot be accessed.
    """
    for proc in psutil.process_iter(['pid', 'name', 'open_files']):
        try:
            if proc.info['name'] == 'EXCEL.EXE':
                for file in proc.info['open_files'] or []:
                    if file.path == file_path:
                        return True
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    return False

def close_workbook(file_path):
    """
    Closes the Excel workbook specified by the file path if it is currently open.

    This function iterates through all running processes to find instances of 'EXCEL.EXE'.
    If the specified file is found to be open by any of these processes, the process is terminated.

    Args:
        file_path (str): The full path to the Excel workbook to be closed.

    Raises:
        psutil.NoSuchProcess: If the process no longer exists.
        psutil.AccessDenied: If the process cannot be accessed due to permission issues.

    Returns:
        None
    """
    for proc in psutil.process_iter(['pid', 'name', 'open_files']):
        try:
            if proc.info['name'] == 'EXCEL.EXE':
                for file in proc.info['open_files'] or []:
                    if file.path == file_path:
                        proc.terminate()
                        proc.wait()
                        return
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

def change_style(self, cell, fill_color=None, font_name=None, font_size=None):
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if font_name or font_size:
        cell.font = Font(name=font_name, size=font_size)


def event_is_multiday(event):
    """
    Check if an event is multi-day.

    Args:
        event (dict): The event dictionary.

    Returns:
        bool:   True if the event is multi-day, 
                False otherwise.
                None if the event does not have a 'date' or 'start_date' and 'end_date' key.
    """
    log.debug(f'Checking if event is multi-day: {event}')
    if 'end_date' and 'start_date' in event.keys():
        log.debug('Event is multi-day')
        return True
    if 'date' in event.keys():
        log.debug('Event is single-day')
        return False
    return None




def iso_year_start_monday(year):
    """
    Returns the Monday of ISO week 1 for the given year.
    (ISO week 1 always includes January 4th.)
    Based on https://en.wikipedia.org/wiki/ISO_8601#Week_dates

    """
    log.debug(f'Calculating ISO year start for {year}')
    jan4 = datetime.date(year, 1, 4)
    return jan4 - datetime.timedelta(days=jan4.isoweekday() - 1)


def get_isoweekday_date_at(target_year=datetime.datetime.now().year, iso_day_of_week=1, mode='first'):
    """
    Get the first or last occurrence of a specific ISO weekday in a given year.
    Parameters:
    -----------
    target_year : int, optional
        The year in which to find the ISO weekday. Defaults to the current year.
    iso_day_of_week : int, optional
        The ISO weekday to find (1 = Monday, 7 = Sunday). Defaults to the ISO8601 default 1 (Monday).
    mode : str, optional
        The mode to determine whether to find the 'first' or 'last' occurrence of the ISO weekday. 
        Must be either 'first' or 'last'. Defaults to 'first'.
    Returns:
    --------
    datetime.date
        The date of the first or last occurrence of the specified ISO weekday in the given year.
    Raises:
    -------
    ValueError
        If `iso_day_of_week` is not between 1 and 7.
        If `mode` is not 'first' or 'last'.
    """
    
    # Validate input
    if iso_day_of_week < 1 or iso_day_of_week > 7:
        raise ValueError("iso_day must be between 1 (Monday) and 7 (Sunday).")
    if mode.lower() not in ['first', 'last']:
        raise ValueError("mode must be either 'first' or 'last'.")
    
    # Get the first day of the year
    first_day_of_year = datetime.date(target_year, 1, 1)
    last_day_of_year = datetime.date(target_year, 12, 31)

    # find the last day of the week
    if mode.lower() == 'last':
        iso_day_of_week = (iso_day_of_week +6) % 7 # Convert to the last day of the week
        # Calculate the difference between the last day of the year and the desired ISO day
        days_until_target_day = (iso_day_of_week + last_day_of_year.isoweekday()) % 7

        # If the last day is already the desired day, return it
        if days_until_target_day == 0:
            return last_day_of_year
        else:
            # Subtract the calculated days to get the previous occurrence of the desired day
            return (last_day_of_year - datetime.timedelta(days=days_until_target_day))

    # find the first day of the week    
    elif mode.lower() == 'first':
        # Calculate the difference between the first day of the year and the desired ISO day
        days_until_target_day = (first_day_of_year.isoweekday() - iso_day_of_week) % 7

        # If the first day is already the desired day, return it
        if days_until_target_day == 0:
            return first_day_of_year
        else:
            # Subtract the calculated days to get the previous occurrence of the desired day
            return (first_day_of_year - datetime.timedelta(days=days_until_target_day))
    
    # If the mode is not recognized, raise an error
    else:
        raise ValueError("mode must be either 'first' or 'last'.")
    
        



def create_date_matrix(year, underflow=0, overflow=0):
    """
    Create a date matrix for a given year with optional underflow and overflow weeks.
    This function generates a nested dictionary structure where the keys are ISO years and weeks,
    and the values are dictionaries containing dates and their corresponding weekdays.
    Args:
        year (int): The year for which to create the date matrix.
        underflow (int, optional): The number of weeks to include before the first week of the year. Defaults to 0.
        overflow (int, optional): The number of weeks to include after the last week of the year. Defaults to 0.
    Returns:
        dict: A nested dictionary where the first level keys are ISO years, the second level keys are ISO weeks,
              and the values are dictionaries containing the date (as ISO format string) and the weekday.
    Example:
        >>> create_date_matrix(2023, underflow=1, overflow=1)
        {
            2022: {52: {datetime.date(2022, 12, 26) 1}, ...},
            2023: {1: {datetime.date(2023, 1, 2): 1}, ...},
            2024: {1:  {datetime.date(2024, 1, 1): 1}, ...}
    """
    # FORCE the ISO8601 week start to Monday - dont accomadate alternative week start days
    
    # Initialize the date matrix
    date_matrix = {}

    # Calculate the start and end dates considering underflow and overflow
    start_date = get_isoweekday_date_at(year, 1, 'first') - datetime.timedelta(weeks=underflow)
    end_date = get_isoweekday_date_at(year, 1, 'last') + datetime.timedelta(weeks=overflow)

    # Iterate through each day from start_date to end_date
    current_date = start_date
    while current_date <= end_date:
        # Get the ISO year and week number
        iso_year, iso_week, iso_weekday = current_date.isocalendar()

        # Create the nested dictionary structure if it doesn't exist
        if iso_year not in date_matrix:
            date_matrix[iso_year] = {}
        if iso_week not in date_matrix[iso_year]:
            date_matrix[iso_year][iso_week] = {}

        # Add the date and weekday to the matrix
        date_matrix[iso_year][iso_week][current_date] = iso_weekday
        
        # iterate - Move to the next day
        current_date += datetime.timedelta(days=1)

    return date_matrix



def set_border(cell, left=None, right=None, top=None, bottom=None, diagonal=None, diagonal_direction=0, outline=None, vertical=None, horizontal=None):
    """
    Set the border of a cell.

    Args:
        cell (openpyxl.cell.cell.Cell): The cell object.
        left (openpyxl.styles.borders.Side): The left border.
        right (openpyxl.styles.borders.Side): The right border.
        top (openpyxl.styles.borders.Side): The top border.
        bottom (openpyxl.styles.borders.Side): The bottom border.
    """
    borders = {}
    if left:
        borders['left'] = left
    if right:
        borders['right'] = right
    if top:        
        borders['top'] = top
    if bottom:
        borders['bottom'] = bottom
    if diagonal:
        borders['diagonal'] = diagonal
    if diagonal_direction:
        borders['diagonal_direction'] = diagonal_direction
    if outline:
        borders['outline'] = outline
    if vertical:
        borders['vertical'] = vertical
    if horizontal:
        borders['horizontal'] = horizontal
    # appply the borders
    cell.border = Border(**borders)


def add_date_event(event_date, event_type, row_index, event_name, event_description, event_dates=event_dates):
    """
    add a new event to a dictionary of dates and return the updated dictionary

    Args:
        event_date (datetime): date of the event
        event_type (str): type of event
        row_index (int): row index for the event
        event_name (str): name of the event
        event_description (str): description of the event
        event_dates (dict): dictionary of events by date

    Returns:
        dictionary: dictionary of events by date with the new event/date added
        
    Structures:
        event_date = {
            xdate: {
                event_type: {
                    row_index: X
                    events: {
                        xeventname: {
                            description: xdescription
        }}}}} 
    
    """  
    # add new date:
    '''
    metadata:
        all events must have either a date, or a start and end date
        all event names of the same type on a given date must be unique
    
    '''
    # check if date exists in event_dates and add it if needed
    if event_date not in event_dates.keys():
        event_dates[event_date] = {}
    
    # check if event_type exists for date and add if needed, add also events, and the row index from row_indicies
    if event_type not in event_dates[event_date].keys():
        event_dates[event_date][event_type] = {
            'row_index': row_index,
            'events': {}
        }
    
    # check if eventname exists, if so raise warning and return event_dates.
    if event_name in event_dates[event_date][event_type]['events'].keys(): 
        log.warning(f'Event {event_name} of type {event_type} already exists for {event_date}. Skipping.')
        return event_dates
    
    # add eventname, and description
    event_dates[event_date][event_type]['events'][event_name]={}
    event_dates[event_date][event_type]['events'][event_name]['description'] = event_description
    
    return event_dates
                
def load_style_config(style_config, initialising_cell, wb=wb, ws=ws):
    """
    Load and apply style configurations to a worksheet.
    Args:
        style_config (dict, optional): A dictionary containing style configurations. 
        initialising_cell (openpyxl.cell.cell.Cell, optional): The cell to apply the initial style to. 
            Defaults to ws['Z999'].
    Returns:
        tuple: A tuple containing:
            - styles (dict): A dictionary of NamedStyle objects keyed by style name.
            - list: A list of style names from the style configuration.
    """
    log.info('Loading style configurations')
    styles={}
    style_names = list(style_config.keys())
    log.debug(f'found {len(style_names)}: {style_names}')
    ## set the initialising cell as the first cell if not provided
    if initialising_cell is None:
        initialising_cell = ws['Z999']

    # iterate over the style configuration 
    for style_name in style_names:
        log.info(f'Processing style: {style_name}')
        new_style = NamedStyle(name=style_name)
        
        # load configuration
        if 'font' in style_config[style_name].keys():
            new_style.font = Font(**style_config[style_name]['font'])
        if 'fill' in style_config[style_name].keys():
            new_style.fill = PatternFill(**style_config[style_name]['fill'])
        
        # borders:
        if 'border' in style_config[style_name].keys():
            border_param = {}
            if 'left' in style_config[style_name]['border'].keys():
                border_param['left'] = Side(**style_config[style_name]['border']['left'])
            if 'right' in style_config[style_name]['border'].keys():
                border_param['right'] = Side(**style_config[style_name]['border']['right'])
            if 'top' in style_config[style_name]['border'].keys():
                border_param['top'] = Side(**style_config[style_name]['border']['top'])   
            if 'bottom' in style_config[style_name]['border'].keys():
                border_param['bottom'] = Side(**style_config[style_name]['border']['bottom'])
            new_style.border = Border(**border_param)

        # alignment and protection
        if 'alignment' in style_config[style_name].keys():
            new_style.alignment = Alignment(**style_config[style_name]['alignment'])
        if 'protection' in style_config[style_name].keys(): 
            new_style.protection = Protection(**style_config[style_name]['protection'])  
        
        # add to ws and initialise.
        styles[style_name] = new_style
        wb.add_named_style(new_style)
        initialising_cell.style = new_style

    # return a dictionary of openpyxl styles and a list of added style names
    return styles, style_names


def populate_column_header(ws, col_start, row_start, column_config, styles, default_heading_style):
    """
    Populate the column headers in a worksheet.
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet object where the headers will be populated.
        col_start (int): The starting column index for the headers.
        row_start (int): The starting row index for the headers.
        column_config (dict): Configuration dictionary for the columns, containing column names and their properties.
        styles (dict): Dictionary of styles to be applied to the headers.
    Returns:
        original column_config with col_index added to each column
    Raises:
        KeyError: If a required key is missing in the configuration dictionary.
    Example:
        populate_column_header(ws, col_start=1, row_start=1, column_config=config['calendar']['columns'], styles=styles)
    """
    
    # populate columns headers
    for i, col in enumerate(column_config):

        # calculate the column index
        column_index = col_start+i # skip the first column to use as labels
        column_config[col]['col_index'] = column_index # save the column index to the config
        log.info(f'Processing column: {col} at index {column_index}')

        # set header style
        # if 'style' in column_config[col].keys():
        #     column_style = column_config[col]['style']
        #     if column_config[col]['style'] in styles.keys():
        #         ws.cell(row=row_start-1, column=column_index, value=col).style = styles[column_style]
        #     else:
        #         log.warning(f"Style {column_style} not found in styles. Using default style.")
        heading_style = 'column_heading'
        ws.cell(row=row_start-1, column=column_index, value=col).style = styles[default_heading_style]
            
            
        # set column width
        if 'width' in column_config[col].keys():
            ws.column_dimensions[get_column_letter(column_index)].width = column_config[col]['width']
        else:
            log.warning(f"Column width not found for {col}. Using default width.")
    
    # return the updated column configuration
    return column_config



def same_year(date1, date2):
    """
    Check if two dates are in the same year.

    Args:
        date1 (datetime.date): The first date to compare.
        date2 (datetime.date): The second date to compare.

    Returns:
        bool: True if both dates are in the same year, False otherwise.
    """
    if not isinstance(date1, datetime.date) or not isinstance(date2, datetime.date):
        raise TypeError("Both date1 and date2 must be of type datetime.date")
    return date1.year == date2.year

def same_month(date1, date2):
    """
    Check if two dates are in the same month.

    Args:
        date1 (datetime.date): The first date to compare.
        date2 (datetime.date): The second date to compare.

    Returns:
        bool: True if both dates are in the same month, False otherwise.
    """
    if not isinstance(date1, datetime.date) or not isinstance(date2, datetime.date):
        raise TypeError("Both date1 and date2 must be of type datetime.date")
    return date1.month == date2.month

def same_day(date1, date2):
    """
    Check if two dates fall on the same day of the month.

    Args:
        date1 (datetime.date): The first date to compare.
        date2 (datetime.date): The second date to compare.

    Returns:
        bool: True if both dates have the same day of the month, False otherwise.
    """
    if not isinstance(date1, datetime.date) or not isinstance(date2, datetime.date):
        raise TypeError("Both date1 and date2 must be of type datetime.date")
    return date1.day == date2.day   

def populate_calendar_dates(ws, date_matrix, row_start, row_count, column_config, styles, border_styles):
    """
    Populates a calendar with dates from a given date matrix into an Excel worksheet.
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet where the calendar will be populated.
        date_matrix (dict): A dictionary containing dates organized by year and week.
        row_start (int): The starting row index for the calendar.
        row_count (int): The number of rows to use for each week.
        column_config (dict): Configuration for the columns, including their indexes and styles.
        styles (dict): A dictionary of styles to apply to the cells.
    Returns:
        dict: A lookup dictionary with dates as keys and their corresponding row and column indexes, week number, and day of the week.

        date_lookup = {
            xdate: {
                row_index: X,
                col_index: Y,
                week: Z,
                weekday: D
            }
        }

    """    
    #set a lookup dictionary
    date_lookup={}
    # unpack border styles
    border_year = Side(**border_styles['year'])
    border_month = Side(**border_styles['month'])
    border_default = Side(**border_styles['default'])
    border_day = Side(**border_styles['days'])

    # assume monday is the start of the week as per ISO8601
    # assume that days are contiguous and that there are no gaps in the dates

    # set initial row and column indexes
    weekdays = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
    weekday_index = 0
    week_start_col = column_config[weekdays[0]]['col_index']
    week_end_col = column_config[weekdays[-1]]['col_index']
    total_weeks = sum(len(weeks) for weeks in date_matrix.values())
    current_row = row_start
    current_col = week_start_col

    # iterate over the date matrix
    for year in date_matrix.keys():
        log.debug(f'adding to calendar year: {year}')
        
        # iterate over the weeks in the year
        for week in date_matrix[year].keys():
            log.debug(f'    adding to calendar week: {week}')

            # label the week column
            _cell_week = ws.cell(row=current_row, column=column_config['week']['col_index'])
            _cell_week.value = week
            _cell_week.style = styles[ column_config['week']['style']]
            for r in range(row_count):
                ws.cell(row=current_row+r, column=column_config['week']['col_index']).style = styles[ column_config['week']['style']]

            # iterate over the dates in the week
            for date in date_matrix[year][week].keys():
                log.debug(f'        adding to calendar date: {date}')
                _cell = ws.cell(row=current_row, column=current_col)

                # calcuate neibouring date cells - None as defualt
                # left cell
                _cell_ = {'left':None, 'right':None, 'top':None, 'bottom':None}
                if current_col > week_start_col:
                    _cell_['left'] = ws.cell(row=current_row, column=current_col-1)
                    
                # top cell
                if current_row > row_start+1:
                    _cell_['top'] = ws.cell(row=current_row-row_count, column=current_col)      

                # set the date, formatting, style
                _cell.value = date
                _cell.style = styles[ column_config[weekdays[weekday_index]]['style'] ]
                _cell.number_format = 'ddd dd-mmm'

                # set all the colum cells border to seperate the week in the cols
                for j in range(col_start, week_start_col):
                    set_border(ws.cell(row=current_row, column=j), top=border_default)


                # add the date to the lookup dictionary
                date_lookup[date] = {
                    'row_index': current_row,
                    'col_index': current_col,
                    'week': week,
                    'weekday': weekday_index+1
                }
                
                # set the borders to seperate months and years
                borders = {'left':border_day, 'right':border_day, 'top':border_day, 'bottom':border_day}
                for position in ['left','top']:
                    if _cell_[position]:
                        if not same_year(date, _cell_[position].value):
                            # add border to seperate years
                            borders[position] = border_year
                        elif not same_month(date, _cell_[position].value): 
                            # add border to seperate months
                            borders[position] = border_month
                set_border(_cell, **borders)

                # set all the other cells in date too
                if borders['left']:
                    for j in range(row_count-1):
                        set_border(ws.cell(row=current_row+j+1, column=current_col), left=borders['left'])
                    

                # iterate days of the week (columns)
                current_col += 1
                weekday_index += 1

                #if current_col > week_end_col:
                #    current_col = week_start_col # reset to the start of the week (- )shouldn't be needed as we are iterating over the date matrix
            
            current_col = week_start_col # reset column to the start of the week
            weekday_index = 0
            
            # iterate weeks (rows)
            current_row += row_count
            
    return date_lookup


def add_column_terms(ws, row_count,  date_lookup, column_config, columns_with_terms, styles):
    """
    Adds terms to specified columns in a worksheet.
    Args:
        ws (Worksheet): The worksheet object where terms will be added.
        row_count (int): The number of rows to increment for each term.
        date_lookup (dict): A dictionary mapping dates to their corresponding row indices.
        column_config (dict): Configuration dictionary for columns, containing column indices.
        columns_with_terms (dict): Dictionary containing terms for each column, with start and end dates and styles.
        styles (dict): Dictionary mapping style names to style objects.
    Returns:
        None
    Raises:
        KeyError: If a required key is missing in the input dictionaries.
        ValueError: If the date lookup fails to find a corresponding row index.
    """

    for col in columns_with_terms.keys():
        log.info(f'Processing column: {col}')
        col_index = column_config[col]['col_index']
                
        for term in columns_with_terms[col]['terms'].keys():
            log.info(f'Processing term: {term}')

            # unpack term details
            start_date = column_config[col]['terms'][term]['start_date']
            end_date = column_config[col]['terms'][term]['end_date']
            # convert to rows:
            start_row_index = date_lookup[start_date]['row_index'] # the start of the rows + offset as the no or rows per week  
            end_row_index = date_lookup[end_date]['row_index'] -1 + row_count # the start of the rows + offset as the no or rows per week
            # unpack style
            style = column_config[col]['terms'][term]['style']

            # iterate over the rows in the column
            current_row = start_row_index
            while current_row <= end_row_index + row_count:
                _cell = ws.cell(row=current_row, column=col_index)
                if 'processed_rox_index' not in column_config[col]['terms'][term].keys():
                    log.debug(f'Adding {term} to {col}')
                    columns_with_terms[col]['terms'][term]['processed_rox_index'] = []
                    _cell.value = term # label the instance of the term

                # add to counter and format cell
                columns_with_terms[col]['terms'][term]['processed_rox_index'] += [current_row]
                _cell.style = styles[ style ]

                # add and end comment
                if (start_row_index != end_row_index) and (current_row == end_row_index + row_count):
                    _cell.value = f'{term} (end)'

                current_row += 1 # increment down rows


    return column_config # pass back the column configuration with the list of row indexes for each term
                


def poulate_events(ws, col_legend, col_date_check, row_start, event_config, dates_lookup, styles, row_count):

    # iterate throguh the event catagories
    for ei, event_cat in enumerate(list(event_config.keys())):
        log.info(f'Adding events from category: {event_cat}')

        # populate the event labels
        current_row = row_start+1
        while ws.cell(row=current_row-1, column=col_date_check).value:
            
            # set the dates label.. even though it will be reset for every category
            #ws.cell(row=current_row-1, column=col_legend).value = 'dates:'
            ws.cell(row=current_row-1, column=col_legend).style = styles[event_config[event_cat]['style_legend_empty']] 
            # alternately use styles['date']                                                                          
        
            # add the event type label
            ws.cell(row=current_row+ei, column=col_legend).value = event_cat
            # use the default (empty) style for the event type label - will be restyled when populated
            ws.cell(row=current_row+ei, column=col_legend).style = styles[event_config[event_cat]['style_legend_empty']]
            # add comment to legend cell
            ws.cell(row=current_row+ei, column=col_legend).comment = Comment(event_config[event_cat]['description'], event_cat)
        
            current_row += row_count
            
        # add event dates
        for eventcount, event in enumerate(event_config[event_cat]['events']):
            # set labels
            event_label_Start = event
            event_label_end = f'{event} (end)'
            log.debug (f'in {event_cat} adding event: {event}')

            # unpack
            start_date = event_config[event_cat]['events'][event]['start_date']
            end_date = start_date
            if 'end_date' in event_config[event_cat]['events'][event].keys():
                end_date = event_config[event_cat]['events'][event]['end_date']
            
            event_days = (end_date - start_date).days + 1
            event_description = event_config[event_cat]['events'][event]['description']

            # check if event is multi-day
            if event_days>1:
                event_label_Start = f'{event} ({event_days} days)'
            event_style = event_config[event_cat]['style']
            legend_style = event_config[event_cat]['style_legend']
            log.debug(f'    {event_label_Start} from {start_date} to {end_date} ({event_days} days)')

            # check dates are in this calendar:
            if start_date not in dates_lookup.keys():
                log.warning(f'Start date {start_date} for event {event} not in calendar')
                if end_date not in dates_lookup.keys():
                    log.warning(f'End date {end_date} for event {event} ALSO not in calendar. Skipping.')
                    continue
                else:
                    log.warning(f'Adjusting start date to first calendar day.')
                    start_date = list(dates_lookup.keys())[0]
                    event_label_Start = f'...{event_label_Start}'

            # add the event name and mark legend column as active
            _cell = ws.cell(row=dates_lookup[start_date]['row_index']+ei+1, column=dates_lookup[start_date]['col_index'])
            ws.cell(row=dates_lookup[start_date]['row_index']+ei+1, column=col_legend).style = styles[legend_style]
            if _cell.value:
                _cell.value = f'{_cell.value}, {event_label_Start}' # concatinate the event name
            else:
                _cell.value = event_label_Start
            # add style
            _cell.style = styles[event_style]

            # add comments
            if _cell.comment:
                if _cell.comment.text:
                    _cell.comment.text += f'\n{event_description}'
            _cell.comment = Comment(event_description, event_cat)
            
            # check to make sure the last day of the event is in the calendar
            if end_date not in dates_lookup.keys():
                # set event end date to the last day of the calendar
                adjusted_end_date = list(dates_lookup.keys())[-1]
                event_days = (adjusted_end_date - start_date).days + 1 # adjust lendth of event for calendar purposes
                event_label_end = f'{event}...'
                log.warning(f'End date {end_date} for event {event} not in calendar. Adjusting end date to last day of calendar.')
            
            # populate subsiquent days of the event
            for d in range(event_days-1):
                _subsiquent_cell = ws.cell(row=dates_lookup[start_date+datetime.timedelta(days=d+1)]['row_index']+ei+1, column=dates_lookup[start_date+datetime.timedelta(days=d+1)]['col_index'])
                _subsiquent_cell_legend = ws.cell(row=dates_lookup[start_date+datetime.timedelta(days=d+1)]['row_index']+ei+1, column=col_legend)
                _subsiquent_cell.style = styles[event_style]
                _subsiquent_cell_legend.style = styles[legend_style]
                
                # add a closing tag on the event end date
                if d==event_days-2:
                    if _subsiquent_cell.value:
                        _subsiquent_cell.value = f'{_subsiquent_cell.value}, {event_label_end}'
                    else:
                        _subsiquent_cell.value = f'{event_label_end}' # mark the end of the event
        
        log.info(f'Added {eventcount+1} events to {event_cat}')


#############################################
log.info('Starting calendar creation')
#############################################

## load configuration
config = load_config(config_file)
log.debug(str(config))

## create calendar
xlsx_file = os.path.join(path, config['calendar']['xlsx_name'])
ws_name = config['calendar']['worksheet_name']
wb = create_calendar(xlsx_file, ws_name, True)
ws = wb[ws_name]
    

# set starting dates
row_start = config['calendar']['start_row']
col_start = config['calendar']['start_column']

# calculate dates:
log.info(f"generating date matrix for {config['calendar']['year']}, underflow:{config['calendar']['underflow']}, overflow:{config['calendar']['overflow']}")
date_matrix = create_date_matrix(year=config['calendar']['year'], underflow=config['calendar']['underflow'], overflow=config['calendar']['overflow'])

# process styles:
styles, style_names = load_style_config(config['styles'], ws['Z999'], wb, ws ) # use A1 as the initialising cell
border_styles = config['side_styles']

# populate column headers and save column index to config
log.info('Populating column headers')
default_heading_style = config['calendar']['column_heading_style']
config['calendar']['columns'] = populate_column_header(ws, col_start, row_start=(row_start+1), column_config=config['calendar']['columns'], styles=styles, default_heading_style=default_heading_style)
columns_with_terms = {c: config['calendar']['columns'][c] for c in config['calendar']['columns'] if 'terms' in config['calendar']['columns'][c]} # get columns with terms

# save row details for later use
row_headings = ['date']+[event for event in list(config['events'].keys())]
row_indicies = {event_type: i for i, event_type in enumerate(row_headings)} #calculate the row (offset) index for each event type
row_count = len(row_headings) # number of rows for each week

# populate the calendar with the date matrix
log.info('Populating calendar with dates')
date_lookup = populate_calendar_dates(ws, date_matrix, row_start=(row_start+1), row_count=row_count, column_config=config['calendar']['columns'], styles=styles, border_styles=border_styles)

# populate the the columns with terms and update the column_config with the row index of the terms
log.info('Populating calendar with terms')
config['calendar']['columns'] = add_column_terms(ws, row_count, date_lookup, config['calendar']['columns'], columns_with_terms, styles)

# populate the label column
log.info('label rows event types')
poulate_events(ws=ws, col_legend=config['calendar']['columns']['legend']['col_index'], col_date_check=config['calendar']['columns']['monday']['col_index'], row_start=row_start+1, event_config=config['events'], dates_lookup=date_lookup, styles=styles, row_count=row_count)         

         

# save the workbook         
save(wb, xlsx_file)
# Open the created Excel file in Excel
if os.name == 'nt':  # Check if the operating system is Windows
    os.startfile(xlsx_file)
else:
    log.error("Opening Excel file is only supported on Windows.")
print ('yo')   
      