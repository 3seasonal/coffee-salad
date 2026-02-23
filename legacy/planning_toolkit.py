import yaml
import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, NamedStyle, Border, Side, Alignment, Protection, Font
import datetime
import os
import logging
import sys
import json


## configuration
# path of this python file
file_path = os.path.abspath(__file__)
config_file_name = 'calendar_config_2025.yaml'
config_file = os.path.join(os.path.dirname(file_path), config_file_name)
    
## globals
# datestamp
ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
ds = datetime.datetime.now().strftime('%Y%m%d')

## logging
# set log file
path = os.path.dirname(file_path)
log_path = os.path.join(path,'logs')
log_file_name = 'calendar_{ts}.log'.format(ts=ts)
log_file = os.path.join(log_path, log_file_name)
# make logs directory if it doesn't exist
if not os.path.exists(log_path):
    os.makedirs(log_path)
    
# configure logging
log = logging.getLogger()
log.setLevel(logging.DEBUG)
logFormatter = logging.Formatter("%(asctime)s [%(threadName)-12.12s] [%(levelname)-5.5s]  %(message)s")
# file output
fileHandler = logging.FileHandler(log_file)
fileHandler.setFormatter(logFormatter)
fileHandler.setLevel(logging.DEBUG)
# console output
consoleHandler = logging.StreamHandler((sys.stdout))
consoleHandler.setFormatter(logFormatter)
consoleHandler.setLevel(logging.INFO)
# add handlers
log.addHandler(consoleHandler)
log.addHandler(fileHandler)

# test logging:
# log.debug('Logging - debug')
# log.info('Logging - info')
# log.warning('Logging - warning')
# log.error('Logging - error')

## global holders
config = {}
wb = None
ws = None
xlsx_file = None
event_dates = {}

def load_config(config_file):
    """
    Load configuration from a YAML file.

    Args:
        config_file (str): The path to the configuration file.

    Globals:
        log (logging.Logger): The logger object.

    Returns:
        dict: The loaded configuration as a dictionary.
    """
    log.info(f'Loading configuration from {config_file}')
    with open(config_file, 'r') as file:
        return yaml.safe_load(file)


def create_calendar(xlsx_file, worksheet_name, overwrite=True):
    """
    Creates a calendar by creating a new workbook, worksheet, and saving it as an Excel file.

    Globals:
        log (logging.Logger): The logger object.

    Returns:
        Workbook: The created workbook object.

    Raises:
        FileExistsError: If the Excel file already exists and overwrite is set to False.
    """
    log.info('Creating calendar')
    
    # create workbook
    log.info(f'Creating calendar {xlsx_file}')
    if os.path.exists(xlsx_file):
        if not overwrite: 
            log.error(f'File {xlsx_file} already exists. Exiting.')
            raise FileExistsError(f'File {xlsx_file} already exists. Overrite set to False. Exiting.') 
        log.warning(f'File {xlsx_file} already exists. Overwriting it.')
    wb = openpyxl.Workbook()

    # configure worksheet
    ws = wb.active
    ws.title = ws_name

    # initialise the first 10 rows and 5 columns
    for x in range(1,201):
       for y in range(1,6):
           ws.cell(row=x, column=y)

    wb.save(xlsx_file)
    return wb

def save(wb=wb, xlsx_file=xlsx_file):
    """
    Save the workbook as an Excel file.

    Args:
        wb (Workbook): The workbook object.
        xlsx_file (str): The path to the Excel file.

    Globals:
        log (logging.Logger): The logger object.
        config (dict): The configuration dictionary.
    
    Raises:
        FileNotFoundError: If the Excel file does not exist.    
    """
    log.info('Saving workbook')
    # if the workbook is not provided, use the global one
    if not xlsx_file:
        xlsx_file = config['calendar']['xlsx_name']
    if os.path.exists(xlsx_file):
        log.info(f'Saving workbook as {xlsx_file}')
        wb.save(xlsx_file)
    else:
        log.error(f'File {xlsx_file} does not exist. Exiting.')
        raise FileNotFoundError(f'File {xlsx_file} does not exist. Exiting.')



def change_style(self, cell, fill_color=None, font_name=None, font_size=None):
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if font_name or font_size:
        cell.font = Font(name=font_name, size=font_size)


def event_is_multiday(event):
    """
    Check if an event is multi-day.

    Args:
        event (dict): The event dictionary.

    Returns:
        bool:   True if the event is multi-day, 
                False otherwise.
                None if the event does not have a 'date' or 'start_date' and 'end_date' key.
    """
    log.debug(f'Checking if event is multi-day: {event}')
    if 'end_date' and 'start_date' in event.keys():
        log.debug('Event is multi-day')
        return True
    if 'date' in event.keys():
        log.debug('Event is single-day')
        return False
    return None




def iso_year_start_monday(year):
    """
    Returns the Monday of ISO week 1 for the given year.
    (ISO week 1 always includes January 4th.)
    Based on https://en.wikipedia.org/wiki/ISO_8601#Week_dates

    """
    log.debug(f'Calculating ISO year start for {year}')
    jan4 = datetime.date(year, 1, 4)
    return jan4 - datetime.timedelta(days=jan4.isoweekday() - 1)

def get_year_weeks_matrix(year, underflow=0, overflow=0):
    """
    Returns a list of dictionaries. Each dictionary looks like:
        {
          "week_no": <int>,  # negative for underflow, 1..(52 or 53) for main year, 1..overflow for overflow
          "dates": [monday_date, tuesday_date, ..., sunday_date]
        }

    :param year: The calendar year you want the ISO weeks for.
    :param underflow: How many extra weeks to include before the start of the ISO year (as negative week numbers).
    :param overflow:  How many extra weeks to include after the end of the ISO year (numbering restarts at 1).
    """
    # Monday of ISO week 1 for this year
    year_start = iso_year_start_monday(year)
    # Monday of ISO week 1 for the next year
    next_year_start = iso_year_start_monday(year + 1)

    # Number of ISO weeks in the specified year
    total_iso_weeks = (next_year_start - year_start).days // 7

    results = []

    # -------------------------
    # 1) Underflow weeks
    #    (labeled as negative)
    # -------------------------
    log.debug(f'calculate underflow weeks: {underflow}')
    for i in range(-underflow, 0):  # e.g. -2, -1 if underflow=2
        monday = year_start + datetime.timedelta(days=7 * i)
        week_dates = [monday + datetime.timedelta(days=d) for d in range(7)]
        results.append({"week_no": i, "dates": week_dates})

    # -------------------------
    # 2) Main year ISO weeks
    #    (labeled 1..52/53)
    # -------------------------
    log.debug(f'calculate main year weeks: {total_iso_weeks}')
    for i in range(1, total_iso_weeks + 1):
        # i-1 so that week 1 starts exactly at year_start
        monday = year_start + datetime.timedelta(days=7 * (i - 1))
        week_dates = [monday + datetime.timedelta(days=d) for d in range(7)]
        results.append({"week_no": i, "dates": week_dates})

    # -------------------------
    # 3) Overflow weeks
    #    (start again at 1..)
    # -------------------------
    ''' update to NOT restart from 1 '''
    log.debug(f'calculate overflow weeks: {overflow}')
    for i in range(total_iso_weeks + 1, total_iso_weeks + 1 + overflow):
        monday = next_year_start + datetime.timedelta(days=7 * (i - 1))
        week_dates = [monday + datetime.timedelta(days=d) for d in range(7)]
        results.append({"week_no": i, "dates": week_dates})
    
    return results


def add_date_event(event_date, event_type, row_index, event_name, event_description, event_dates=event_dates):
    """
    add a new event to a dictionary of dates and return the updated dictionary

    Args:
        event_date (datetime): date of the event
        event_type (str): type of event
        row_index (int): row index for the event
        event_name (str): name of the event
        event_description (str): description of the event
        event_dates (dict): dictionary of events by date

    Returns:
        dictionary: dictionary of events by date with the new event/date added
        
    Structures:
        event_date = {
            xdate: {
                event_type: {
                    row_index: X
                    events: {
                        xeventname: {
                            description: xdescription
        }}}}} 
    
    """  
    # add new date:
    '''
    metadata:
        all events must have either a date, or a start and end date
        all event names of the same type on a given date must be unique
    
    '''
    # check if date exists in event_dates and add it if needed
    if event_date not in event_dates.keys():
        event_dates[event_date] = {}
    
    # check if event_type exists for date and add if needed, add also events, and the row index from row_indicies
    if event_type not in event_dates[event_date].keys():
        event_dates[event_date][event_type] = {
            'row_index': row_index,
            'events': {}
        }
    
    # check if eventname exists, if so raise warning and return event_dates.
    if event_name in event_dates[event_date][event_type]['events'].keys(): 
        log.warning(f'Event {event_name} of type {event_type} already exists for {event_date}. Skipping.')
        return event_dates
    
    # add eventname, and description
    event_dates[event_date][event_type]['events'][event_name]={}
    event_dates[event_date][event_type]['events'][event_name]['description'] = event_description
    
    return event_dates
                


#############################################
log.info('Starting calendar creation')
#############################################

## load configuration
config = load_config(config_file)
log.debug(str(config))

## create calendar
xlsx_file = os.path.join(path, config['calendar']['xlsx_name'])
ws_name = config['calendar']['worksheet_name']
wb = create_calendar(xlsx_file, ws_name, True)
ws = wb[ws_name]
    

# set starting dates
row_start = config['calendar']['start_row']
col_start = config['calendar']['start_column']

# calculate dates:
log.info(f"generating date matrix for {config['calendar']['year']}, underflow:{config['calendar']['underflow']}, overflow:{config['calendar']['overflow']}")
date_matrix = get_year_weeks_matrix(year=config['calendar']['year'], underflow=config['calendar']['underflow'], overflow=config['calendar']['overflow'])
log.debug(f'Date matrix: {date_matrix}')

# process styles:
style_config = config['styles']
style_names = list(style_config.keys())
styles = {}
for style_name in style_names:
    log.debug(f'Processing style: {style_name}')
    new_style = NamedStyle(name=style_name)
    
    # load configuration
    new_style.font = Font(**style_config[style_name]['font'])
    new_style.fill = PatternFill(**style_config[style_name]['fill'])
    new_style.border = Border(
        left=Side(**style_config[style_name]['border']['left']),
        right=Side(**style_config[style_name]['border']['right']),
        top=Side(**style_config[style_name]['border']['top']),
        bottom=Side(**style_config[style_name]['border']['bottom'])
    )
    new_style.alignment = Alignment(**style_config[style_name]['alignment'])
    new_style.protection = Protection(**style_config[style_name]['protection'])  
    
    # add to ws and initialise.
    styles[style_name] = new_style
    wb.add_named_style(new_style)
    ws['A1'].style = new_style
    
    
# populate columns headers
for i, col in enumerate(config['calendar']['columns']):

    # add extra column for labels
    config['calendar']['columns'][col]['col_index'] = col_start+i #skip the first column to use as labels
    ws.cell(row=row_start, column=col_start+i, value=col).style = styles[style_names[0]]
    
    # where there are 'terms' - or things with start and end dates check:
    # if weeks list does not exist, add:
    # lookup start date in matrix, add week no to weekslist
    # lookup end date in matrix, add week no to weekslist
    # calcuate all the weeks inbetween week at index 0 and 1, add to list.
    
log.info(f"Column headings: {list(config['calendar']['columns'])}")

# process column headings by date - if there is a term make a subset
columns_with_terms={}
for c in config['calendar']['columns'].keys():
    if 'terms' in config['calendar']['columns'][c].keys():
        columns_with_terms[c] = config['calendar']['columns'][c]
        

        
        
        
                
# save row headings and indicies for later use
log.debug(f'Row headings from config: {config["events"].keys()}')
row_headings = ['date']+[event for event in list(config['events'].keys())]
row_indicies = {}

#calculate the row (offset) index
for i, event_type in enumerate(row_headings):
    row_indicies[event_type]=i
log.debug(f'Row headings: {row_headings}')

# set start row
row_count = len(row_headings)


current_row = row_start +1 #skip the header row

# iterate over the date matrix
for week in date_matrix:
    week_no = week['week_no']
    log.info(f'Processing week: {week_no}')
    
    # populate week number
    ws.cell(row=current_row, column=config['calendar']['columns']['week']['col_index'], value=week_no)

    # populate row legend column
    log.debug(f'addining legend items {row_headings}')
    for i, rh in enumerate(row_headings):
        ws.cell(row=current_row+i, column=config['calendar']['columns']['legend']['col_index'], value=rh)

    # polulate terms
    '''
    use - columns_with_terms
    return max(start1, start2) <= min(end1, end2)
    
    '''
    

    # Calculate events contents
    ##########################
    #########################

    '''UPDATE ALL ITERATORS such that don't need to use a pre-decrement'''
    current_col = config['calendar']['columns']['Monday']['col_index'] - 1
    
    # iterate days of the week
    for date in week['dates']:
        current_col += 1
        
        # add date
        log.debug(f'Date: {date}')
        date_formatted = date.strftime('%b-%d')
        ws.cell(row=current_row, column=current_col, value=date_formatted)
        
        # add event content
        if date in event_dates.keys():
            for event_type in event_dates[date]:
                
                # calculate event attributes
                target_row = current_row + event_dates[date][event_type]['row_index']
                event_names = list(event_dates[date][event_type]['events'].keys())
                log.debug(f'same day events for {event_type} at {date} - {event_names}')
                event_name=','.join(event_names) # concatinate multiple events into a single event
                # use dectiption as a hoverover
                log.debug(f"event date {date}: {event_dates[date]}")
                
                event_description= ','.join([event_dates[date][event_type]['events'][event_name]['description'] for event_name in event_names]) # concatinate multiple events into a single event
                # add event name and description as a comment.
                ws.cell(row=target_row, column=current_col, value=event_name)
                comment = Comment(event_description, event_name)
                ws.cell(row=target_row, column=current_col).comment = comment
    
    # increment the rows to the next week
    row_start += row_count

    ##########################
    # populate events
    # build date based lookup dictionary of events against dates
    
    for event_type in config['events'].keys():
        for event in config['events'][event_type].keys():
            
            #unpack row index
            event_row_index = row_indicies[event_type]
            if event_is_multiday(config['events'][event_type][event]):
                
                # validate the event:
                if ('start_date' not in config['events'][event_type][event].keys()) or ('end_date' not in config['events'][event_type][event].keys()):
                    # invalid data - start or end date not found in event
                    log.error(f'Invalid data: Start or end date not found for event {event} of type {event_type}. BREAKING.')
                    raise ValueError(f'Invalid data: Start or end date not found for event {event} of type {event_type}.')

                # create an entry for each continuous day
                current_date = config['events'][event_type][event]['start_date']
                while current_date <= config['events'][event_type][event]['end_date']:
                    
                    # add event date:
                    event_dates = add_date_event( current_date, event_type, event_row_index, event, config['events'][event_type][event]['description'], event_dates)
                    # increment day (of multi-day event) by one day:
                    current_date += datetime.timedelta(days=1)    
            
            else:
                # add event as a single day event
                if 'date' not in config['events'][event_type][event].keys():
                    log.error(f'Invalid data: Date not found for event {event} of type {event_type}. BREAKING.')
                    raise ValueError(f'Invalid data: Date not found for event {event} of type {event_type}.')
                event_dates = add_date_event( config['events'][event_type][event]['date'], event_type, event_row_index, event, config['events'][event_type][event]['description'], event_dates)
    
    current_row = current_row + row_count
                    
save(wb, xlsx_file)
    
      