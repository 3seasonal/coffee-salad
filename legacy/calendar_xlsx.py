from openpyxl import Workbook 
from openpyxl import worksheet
import os

class calendar_xlsx:
    """
    This class is used to manage the xlsx file



    """
    def __init__(self, target_xlsx_file=None, target_worksheet=None, utility_class=None):
        """Constructor - Initialize the object"""
        self._wb = None
        if target_xlsx_file is not None:
            if os.path.exists(target_xlsx_file):
                self._wb = Workbook(target_xlsx_file)
        
        # set the utility class to an instatiated object
        self._cu = utility_class if utility_class is not None else None
        print (f'logging enabled in calendar_xlsx: {self.logging_enabled()}')

        # report on workbook:
        if self._wb is not None:
            if self._cu is not None:
                self._cu.log(f'Workbook {target_xlsx_file} opened')
                self._cu.log(f'contains {len(self._wb.sheetnames)} worksheets: {self._wb.sheetnames}')
            else:
                print(f'Workbook {target_xlsx_file} opened')
                print(f'contains {len(self._wb.sheetnames)} worksheets: {self._wb.sheetnames}')

    def logging_enabled(self):
        """Check if logging is enabled"""
        return self._cu is not None
    
    
    # get the list of columns as defined in the worksheets metadata
    def get_column_list_from_ws(self, source_worksheet=None):
        
        # error handling
        if self._wb is None:
            self._cu.log('No workbook opened. Exiting.')
            raise ValueError('No workbook opened. Exiting.')

        if source_worksheet is None:
            self._cu.log('worksheet {source_worksheet} not found')
            raise ValueError(f'worksheet {source_worksheet} not found.')
        
        # start
        self._cu.log(f'get column list from worksheet {source_worksheet}')
        ws = self._wb[source_worksheet]
              
        # read the first row of the worksheet, does it have a value column?
        value_col_index = get_column_index(self, ws, 'value')
        if value_col_index is None:
            self._cu.log(f'"value" column not found in the first row of worksheet {source_worksheet}')
            raise ValueError(f'"value" column not found in the first row of worksheet {source_worksheet}')
        
        # check the first row of the worksheet
        first_row = [cell.value.lower() for cell in source_worksheet[1]]
        second_row = [cell.value.lower() for cell in source_worksheet[2]]
        
        # Check if the first row has the expected format
        if first_row[:3] != ['param', 'type', 'value']:
            self._cu.log(f'formatting error in configuration worksheet {source_worksheet} first row is not: param, type, value')
            raise ValueError(f'formatting error in configuration worksheet {source_worksheet} first row is not: param, type, value')
    
        column_headers = 
        
                
    
        self._cu.log(f'"value" column not found in the first row of worksheet {source_worksheet}')
        raise ValueError(f'"value" column not found in the first row of worksheet {source_worksheet}')
        
        # Get the index of the "value" column
        get_column_index(self, source_worksheet=None, search_value='value'):value_column_index = first_row.index("value")
        self._cu.log(f'"value" column index in the first row of worksheet {source_worksheet}: {value_column_index}')
        
    
        self._cu.log(f'"value" column found in the first row of worksheet {source_worksheet}')
        
        # check the second column
        
        # check for the "value" column
        
        #create a list of columns - first three are param, type - the remainder are the columns in the comma seperated list in the value column
        
        
    # get the index of the value column:
    def get_column_index(self, source_worksheet=None, search_value='value'):
        first_row = [cell.value.lower() for cell in source_worksheet[1]]
        # Check if "value" column exists
        if search_value not in first_row:
            return None
        return first_row.index(search_value)
        
        
        
    
    # recursive function that
    # reads a sheet, returns a dictionary object of the sheet
    def get_config_from_xlsx(self, source_xlsx_file=None, source_worksheet=None, is_styles=False):
        """Get the configuration"""
        if source_xlsx_file is not None:
            if not os.path.exists(source_xlsx_file):
                raise FileNotFoundError(f'File {source_xlsx_file} does not exist. Exiting.')
            else:
                source_wb = Workbook(source_xlsx_file)
                if source_worksheet is None:
                    raise ValueError('No worksheet specified. Exiting.')
                else:                    
                    if source_worksheet not in source_wb.sheetnames:
                        raise ValueError(f'Sheet {source_worksheet} not found in {source_xlsx_file}. Exiting.')
                    else:
                        source_ws = source_wb[source_worksheet]

                        #### parse worksheet
                    
                        '''
                        RECURSIVE - ADJUST TO RETURN NONE instead of worksheets - log errors to log file
                        build dictionary of sheet
                            while elements in sheet
                            handle formats?
                            identify links Call recursively and assign to dictionary element

                            if styles:
                            - process with openpyxl

                            return dictionary


                        meta params:
                         - flag meta columns
                                                  - 

                        '''
