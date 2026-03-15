#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font  
import datetime
from typing import Dict, List, Any, Tuple, Optional, Union
import os
import logging
import psutil


"""Calendar xlsx writer module
    Coffee-Salad calendar package

    This module provides cordination functions for calendar creation.
    It maintains creates and or updates calendars using xlsx files.
    
    Author:     dillonj
    Created:    2025-08-29
    Description: class for writing xlsx files .

"""

    
class CalendarXlsxCreatorError():
    def __init__(self, raised_error_message: str):
        '''
        Simple error class for handling errors in the calendarXlsxCreator class. It takes a string message as input and formats it for display.
        
        Args:
            raised_error_message (str): The error message to be displayed when the exception is raised.
        Returns:
            None
        Raises:
            None
        '''
        super().__init__(
            f"Exception in the XLSX Creator: {raised_error_message}"
        )
        

class calendarXlsxCreator:
    """
    Creates a XLSX file from scratch
    See calendarXLSXUpdator for updating an existing XLSX file.
    """
    

    
    def __init__(self, config: dict, logger=None, ts=datetime.datetime.now().strftime('%Y%m%d_%H%M%S')):
        """Initialize the xlsx calenadarcreator with the path to the Excel file.
        calculated relevant lookup values

        Args:
            config (dict): The configuration dictionary containing calendar settings.
            logger (logging.Logger, optional): A logger instance for logging messages. Defaults to None.
            ts (str, optional): A timestamp string to be used in the file name. Defaults to the current date and time in 'YYYYMMDD_HHMMSS' format.
            
        Returns:
            None
            
        Raises:
            CalendarXlsxCreatorError: If there is an error during initialization, such as missing configuration.
        
        """
        
        #set up logging
        self.log = logger or logging.getLogger(__name__)
        
        # set output path
        self.script_dir = os.path.dirname(os.path.abspath(__file__))

        #get general unitlity class:
        dt = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Store configuration for use:
        if not config:
            self.log.error("No configuration provided to calendarXlsxCreator.")
            raise CalendarXlsxCreatorError("Oh no. The XLSX writer was not passed a configuration. :(")
        else:
            self.config = config # store the config for later use, but it is not used directly in this class. It is passed to the calendarXLSXUpdator for processing.
            # contains sub-dictionaries: calendar, columns, events, styles.
                   
        # Handle the creation of the calendar XLSX file. If it does not exist,
        self.xlsx_name = config.get("calendar", {}).get("xlsx_name")
        if not self.xlsx_name.endswith(".xlsx"):
            self.xlsx_name = self.xlsx_name + ".xlsx"
        self.xlsx_name = self.xlsx_name.replace(".xlsx", f"_{dt}.xlsx") # add time stamp to the file name 
    
        # calculate path to the xlsx file:
        self.xlsx_path = os.path.join(self.script_dir, self.xlsx_name)
        self.log.info(f"XLSX file path set to: {self.xlsx_path}")    
        
        # initialise configuration handlers:
        self.events = config['events'] # the events configuration, used for populating the calendar with events
        self.category_list = list(self.events.keys()) # list of categories in the calendar, used for indexing the columns 
        self.calender_config = config['calendar']

        # upack and store the styles from the configuration for later use:
        self.log.debug("unpack style configuration from config")
        self.style_config = config['styles']

        # create new xlsx file using openpyxl (check if it exists first, if it does, delete it and create a new one)
        if os.path.exists(self.xlsx_path):
            self.log.warning(f"XLSX file already exists at {self.xlsx_path}. It will be overwritten.")
            os.remove(self.xlsx_path)
        self.workbook = openpyxl.Workbook()
        self.log.info(f"New XLSX workbook created at {self.xlsx_path}.")
        self.worksheet = self.workbook.active
        self.worksheet.title = (self.calender_config['worksheet_name'])
        self.cell_date={} # a dictionary to store the cell references for each date

        # get key dates 
        self.start_date = datetime.strptime(self.calender_config['start_date'], "%Y-%m-%d").date()
        self.end_date = datetime.strptime(self.calender_config['end_date'], "%Y-%m-%d").date()
        underflow_delta = datetime.timedelta(weeks = self.calender_config['underflow_weeks'])
        overflow_delta = datetime.timedelta(weeks = self.calender_config['overflow_weeks'])
        self.first_date = self.start_date - underflow_delta
        self.last_date = self.end_date + overflow_delta

        # check if the start date is the correct day of the week as defined in the config (week_stats_on). 
        # if it is not, change the start date to the previous day until it is the correct day of the week. This is to ensure that the calendar starts on the correct day of the week and that the dates are aligned correctly in the calendar.
        self.start_isoweekday = self.calender_config['week_stats_on']
        while self.start_date.isoweekday() != self.start_isoweekday:
            self.start_date = self.start_date - datetime.timedelta(days=1)
        
        # check if the end date is the correct day of the week as defined in the config (week_stats_on -1).where the number is between 1 and 7
        self.end_isoweekday = ((self.calender_config['week_stats_on'] - 1) % 7 or 7) # this calculation ensures that if week_stats_on is 1 (Monday), the end date should be Sunday (7), and if week_stats_on is 7 (Sunday), the end date should be Saturday (6).
        while self.end_date.isoweekday() != self.end_isoweekday:
            self.end_date = self.end_date + datetime.timedelta(days=1)

        # get total weeks
        self.total_weeks = int( ((self.last_date - self.first_date).days + 1) // 7 )
        self.log.info(f"Calendar will cover the date range from {self.first_date} to {self.last_date}, which is a total of {self.total_weeks} weeks.")

        # get column configuration:
        self.column_config = self.config['columns']
        dow_cols = self.column_config['days_of_week_columns']
        self.column_list = self.column_config['column_order']
        self.first_dow_column = self.column_list.index(dow_cols[0]) # get the column index of the first day of week column 
        self.column_list.pop(self.first_dow_column) # remove the day of week columns from the column list, as they will be inserted later in the correct order based on the start_isoweekday

        # isoweekday lookup
        self.isoweekday_name={ 1: "Monday", 2: "Tuesday", 3: "Wednesday", 4: "Thursday", 5: "Friday", 6: "Saturday", 7: "Sunday" }
        self.weekday_names = [v.lower() for v in self.isoweekday_name.values()]
        # insert the weekday names into the column_list,
        iwd = self.start_isoweekday
        index = 0
        while index < 7:
            self.column_list.insert(self.first_dow_column + index, self.isoweekday_name[iwd])
            iwd += 1 if iwd < 7 else 1
            index += 1

        # confirm the start_column and start_row are valid (+ve ints and not zero):
        self.calendar_start_coloumn = self.calender_config['start_column']
        if not isinstance(self.calender_config['start_column'], int) or self.calender_config['start_column'] <= 0:
            self.log.critical(f"Invalid start_column value: {self.calender_config['start_column']}. DEFAUTING to 1.")
            self.calendar_start_coloumn = 1
        self.calendar_start_row = self.calender_config['start_row']
        if not isinstance(self.calender_config['start_row'], int) or self.calender_config['start_row'] <= 0:
            self.log.critical(f"Invalid start_row value: {self.calender_config['start_row']}. DEFAUTING to 1.")
            self.calendar_start_row = 1

        # calculate row offsets, excluding header row
        self.row_offsets = len(self.category_list)+1 #for the date row.
        

        # calculate date cell references and store in the cell_date dictionary:
        self.cell_date = {} # a dictionary to store the date of a cell references
        self.date_cell = {} # a dictionary to store the cell reference of a date
        self.max_row = 1 # stored the maxium row number
        self.max_col = 1 # stored the maxium column number
        row = self.calender_config['start_row']+1 # start from the row below the header row
        col = self.first_dow_column
        week = 0
        dayno = 0
        date = self.first_date
        while week < self.total_weeks:
            while dayno < 7:
                self.cell_date[(row,col)] = date
                self.date_cell[date] = (row, col)
                dayno += 1
                col += 1
                date += datetime.timedelta(days=1)
                self.max_row = max(self.max_row, row)
                self.max_col = max(self.max_col, col)
            week += 1
            dayno = 0
            row += self.row_offsets
            col = self.first_dow_column
        
        # initialise cells for later use:
        for x in range(1,self.max_col):
            for y in range(1, self.max_row):
                self.worksheet.cell(row=x, column=y)

        # done
        self.log.info(f"Initialised calender writer")
        self.save()



    def category_row_offset(self, category: str) -> int:
        """Calculate the row offset for a given category based on the category list.
        
        Args:
            category (str): The name of the category for which to calculate the row offset.

        Returns:
            int: The row offset for the given category, which is the index of the category in the category list plus one (to account for date row).
        
        Raises:
            CalendarXlsxCreatorError: If the category is not found in the category list.
        """
        if category not in self.category_list:
            self.log.error(f"Category '{category}' not found in category list.")
            raise CalendarXlsxCreatorError(f"Category '{category}' not found in category list.")
        return self.category_list.index(category) + 1 # add 1 to account for date row

        

        
    def save(self):
        """Save the workbook to the specified path.
        
        Args:
            None
        Retrns:
            None
        
        Raises:
            CalendarXlsxCreatorError: If there is an error saving the workbook.
        
        """
        try:
            self.workbook.save(self.xlsx_path)
            self.log.info(f"Workbook saved successfully at {self.xlsx_path}.")
        except Exception as e:
            self.log.error(f"Failed to save workbook: {e}")
            raise CalendarXlsxCreatorError(f"Failed to save workbook: {e}")




    def mark_date_as_busy(self, date: datetime.date):
        """Mark a specific date as busy with an event in the calendar by changing its style.
        
        Args:
            date (datetime.date): The date to be marked as busy.
            category (str): The category of the event (used for column indexing).
            event_name (str): The name of the event to be displayed in the cell.
        
        Returns:
            Bool: True if the date style was updated, False if it was already marked as busy.
        
        Raises:
            CalendarXlsxCreatorError: If there is an error marking the date as busy, such as invalid date or category.
        
        """
        # This function will be implemented later. It will use the trackers to find the correct cell based on the date and category, and then update the cell with the event name and appropriate styling.
        pass



    def _apply_cell_style(self, cell, style_name, border=False):
        """Apply a style to a cell based on the style configuration. 
            See https://openpyxl.readthedocs.io/en/3.1/styles.html
            Note that named styles are not used as they are mutible and later style changes will not affect output.

        Args:
            None

        Returns:

        """
        pass

    def _build_calendar_structure(self):
        """Build the basic structure of the calendar worksheet.
        
        Args:
            None

        Returns:
            None
        """
        # populate the header row
        self.log.debug("populate header row")
        for column_no, header in enumerate(self.column_list):
            col = column_no + self.calendar_start_coloumn
            self.worksheet.cell(row=self.calendar_start_row, column=col).value = header
            # apply style to the header row ... tbc

        # populate date cells:
        self.log.debug("populate dates")
        self.weekno_column = self.column_list.index(self.column_config['weekno_column'])+ self.calendar_start_coloumn
        for date in self.date_cell:
            cell = self.worksheet.cell(row=self.date_cell[date][0], column=self.date_cell[date][1])
            cell.value = date
            cell.number_format = 'ddd dd-mmm'
            # calculate the week no column. if this is the first day of the week, as defined in the config (week_stats_on)
            if date.isoweekday() == self.self.start_isoweekday:
                self.worksheet.cell(row=self.date_cell[date][0], column=self.weekno_column).value = date.isocalendar()[1] 
            # apply style to the date and catagory cells


        # configure columns

        # iterate the columns list
        for i, column_name in enumerate (self.column_list):
            column_name = column_name.lower()
            col = i + self.calendar_start_coloumn


            #FETCH CONFIG
            (xxxxxxxxxxxxxxxx)

        
            # process weekno
            if column_name.lower() in ('weekno','week'):
                self.log.debug("populate legend column")
                self.legend_column = self.column_list.index(self.column_config['legend_column']) + self.calendar_start_coloumn
                colno = self.legend_column
                rowno = self.calendar_start_row + 1 #skip over the column header row
                while rowno < self.max_row:
                    rowno +=1 #skip over the date row
                    for cr, catagory in enumerate(self.category_list):
                        self.worksheet.cell(row=rowno+cr, column=colno).value = catagory
                        # apply style to the catagory column ... tbc

            # process uni
            if column_name.lower() in ('uni'):
        
            # process school
            if column_name.lower() in ('school'):
        
            # legend
            if column_name.lower() in ('legend'):

            # weekday
            if column_name.lower() in self.weekday_names:
        
            # notes
            if column_name.lower() in ('notes'):
        
            # leave
            if column_name.lower() in ('leave'):




        # populate catagory columns:
        
        



    def process_catagories(self):
        """Process the catagories from the configuration and populate the legend column with the catagory names.
        
        Args:
            None

        Returns:
            None
        """
        pass


    def process_events(self):
        """Process the events from the configuration and populate the calendar with the event data.
        
        Args:
            None

        """
        pass
        
        
        # iterate events by catagory type and populate the cells with the event data.
        
        # if the event is a multi-day event, populate the cells for each day of the event duration, and apply the appropriate styling to indicate that it is a multi-day event. This may involve merging cells or applying a specific style to the range of cells that the event spans.
        # if there is already content in the cell, append to the existing

        #style the legend column appropriately

        # colour approparietately based on the event category and the styles defined in the configuration



## algorithm - to be implemetned later:

'''
create worksheet

save


functions:
- load config (on initialisation)
- create worksheet
- save workbook



- update cell style
- add style to workbook

- event_is_multiday

- add event.

---
create matrix.
    create trackers:
        catagory list - index of each catagory releative to the date row
        date index - given a date, return the row and col index
        days of the week 2 column - given a day of the week which column
        column 2 day of the week - given a column, return the day of the week index.
        day of the weeek - iso index
        list of styles created
        
        SELF.
        
    create styles
        and add to style created list
        
    populate numbers and dates in calendar.
        and colour the cells as they are created.
    
    
iterate events and populate the cells with the event data.
    colour approparietately
    
    event matrix list - a dictionary of of key dates, 
            each with a list of catagories.
                each catagory with a list of events. (that are the key wihtin the config.events)
                    each event is a touple of (event name, #day, of #days)
    (in this way can events for colouring purposes)
    
    
Add event

    



'''

    
def is_workbook_open(file_path):
    """
    Check if a specific Excel workbook is currently open.

    Args:
        file_path (str): The full path to the Excel workbook file.

    Returns:
        bool: True if the workbook is open, False otherwise.

    Raises:
        psutil.NoSuchProcess: If the process no longer exists.
        psutil.AccessDenied: If the process cannot be accessed.
    """
    for proc in psutil.process_iter(['pid', 'name', 'open_files']):
        try:
            if proc.info['name'] == 'EXCEL.EXE':
                for file in proc.info['open_files'] or []:
                    if file.path == file_path:
                        return True
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    return False

def close_workbook(file_path):
    """
    Closes the Excel workbook specified by the file path if it is currently open.

    This function iterates through all running processes to find instances of 'EXCEL.EXE'.
    If the specified file is found to be open by any of these processes, the process is terminated.

    Args:
        file_path (str): The full path to the Excel workbook to be closed.

    Raises:
        psutil.NoSuchProcess: If the process no longer exists.
        psutil.AccessDenied: If the process cannot be accessed due to permission issues.

    Returns:
        None
    """
    for proc in psutil.process_iter(['pid', 'name', 'open_files']):
        try:
            if proc.info['name'] == 'EXCEL.EXE':
                for file in proc.info['open_files'] or []:
                    if file.path == file_path:
                        proc.terminate()
                        proc.wait()
                        return
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue