

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
from typing import Dict, List, Any, Tuple, Optional, Union
import os
import logging

# Configure logging
logger = logging.getLogger("ConfigXlsxReader")
logger.setLevel(logging.DEBUG)

# Console handler for info-level logs
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter("%(levelname)s: %(message)s")
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)

# File handler for debug-level logs
file_handler = logging.FileHandler("config_xlsx_reader.log", mode="a")
file_handler.setLevel(logging.DEBUG)
file_formatter = logging.Formatter("%(asctime)s %(levelname)s [%(name)s]: %(message)s")
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

# Custom exception classes for ConfigXlsxReader
class ConfigXlsxReaderError(Exception):
    """Base exception for ConfigXlsxReader errors."""
    pass

class MissingColumnError(ConfigXlsxReaderError):
    def __init__(self, column: str, sheet: str):
        super().__init__(f"Missing required column '{column}' in sheet '{sheet}'.")

class InvalidTypeError(ConfigXlsxReaderError):
    def __init__(self, param: str, expected: str, actual: str, sheet: str, row: int):
        super().__init__(
            f"Invalid type for parameter '{param}' in sheet '{sheet}', row {row}: expected {expected}, got '{actual}'."
        )

class ConfigReferenceError(ConfigXlsxReaderError):
    def __init__(self, param: str, sheet: str, row: int):
        super().__init__(
            f"Empty config reference for parameter '{param}' in sheet '{sheet}', row {row}."
        )

class ConfigXlsxReader:
    """
    Reads calendar configuration from an Excel file.
    Extracts configuration into four dictionaries:
    - calendar: Calendar configuration (start/end dates, etc.)
    - columns: Calendar column structures
    - events: Event types and entries
    - styles: Styles used in the calendar
    """
    
    def __init__(self, xlsx_path: str):
        """Initialize the reader with the path to the Excel file."""
        self.xlsx_path = xlsx_path
        self.worksheet = {} #list of worksheets in the workbook
        
        # Output dictionaries
        self.calendar = {}
        self.columns = {}
        self.events = {}
        self.styles = {}       
       
        # check if the xlsx file exists
        if not self.xlsx_path:
            raise ConfigXlsxReaderError("Excel file was not provided.")
        if not os.path.isfile(self.xlsx_path):
            raise ConfigXlsxReaderError(f"Excel file '{self.xlsx_path}' does not exist.")
        
        # check if it contains a valid config-main sheet
        # Load workbook in read-only mode
        self.workbook = openpyxl.load_workbook(self.xlsx_path, read_only=True)
        if "config-main" not in self.workbook.sheetnames:
            self.workbook.close()
            raise MissingColumnError("config-main", self.xlsx_path)
        
        else:
            # report the number of sheets loaded
            logger.info(f"Loaded workbook '{self.xlsx_path}' with {len(self.workbook.sheetnames)} sheets.")
            logger.debug(f"Workbook sheets: {self.workbook.sheetnames}")
         
    
    def read_config(self):
        """
        Reads the configuration from the Excel file and returns the four dictionaries.
        Follows the configuration tree, starting from config-meta, then config-main, and only parses referenced config sheets (ignoring config-meta as a config).
        Returns:
            Tuple containing (calendar, columns, events, styles) dictionaries
        """       
        # parse the config-main sheet and follow the configuration tree
        parsed_config = self._parse_worksheet_("config-main")

        return parsed_config['calendar'], parsed_config['columns'], parsed_config['events'], parsed_config['styles']



    def _parse_worksheet_(self, worksheet_name):
        """_summary_
        This function parses a single worksheet from the configuriation and returns 
        a dictionary representation of its contents.
        Function is recursive and will follow config references to other sheets.
        will check the first two rows for meta-config and header row.
        This will be used to configure how the rest of the sheet is parsed.
        Args:
            worksheet_name (string): the name of the worksheet to parse
        Returns:
            Dict: a dictionary representation of the worksheet contents
        THrows:
            ValueError: if the worksheet does not exist
        """

        
        # get the worksheet and meta-config
        worksheet, col_index, col_value_names = self._get_worksheet_metaconfig_(worksheet_name)
        
        # iniatialize output dictionary
        output_dict = {}
        

        # Build a dictionary mapping column names to their indices
        col_names = {name.strip().lower(): col_index.get(name) for name in (['param', 'type'] + col_value_names)}


        # Parse the content in the sheet based on the meta-config
        for row in worksheet.iter_rows(min_row=3):
            #if the param cell is empty or commented out, skip the row
            if (row[col_names['param'] - 1].value is not None) and (row[col_names['param'] - 1].value.strip() != "") and not (row[col_names['param'] - 1].value.strip().startswith("#")):
                param = row[col_names['param'] - 1].value.strip().lower()
                param_type = row[col_names['type'] - 1].value.strip().lower() if row[col_names['type'] - 1].value else "string"
                value = row[col_names['value'] - 1].value


                # Handle different prinitive types for the row value
                if param_type.startswith("style-cell"):
                    output_dict[param] = self._extract_cell_style_(row[col_names['value'] - 1])

                elif param_type.startswith("style-border"):
                    output_dict[param] = self._extract_border_style_(row[col_names['value'] - 1])

                elif param_type == "string":
                    output_dict[param] = value

                elif param_type == "bool":
                    output_dict[param] = True if value in ["true","True","TRUE","yes","Yes","YES","1"] else False

                elif param_type == "int":
                    try:
                        output_dict[param] = int(value)
                    except (ValueError, TypeError):
                        raise InvalidTypeError(param, "int", value, worksheet_name, row[0].row)

                elif param_type == "float":
                    try:
                        output_dict[param] = float(value)
                    except (ValueError, TypeError):
                        raise InvalidTypeError(param, "float", value, worksheet_name, row[0].row)
                    output_dict[param] = float(value)

                elif param_type == "date":
                    output_dict[param] = self._parse_date_(value)


                # Handle different complex types
                elif param_type == "list":

                    # create a sub-dictionary for the list
                    sub_dict = {}

                    for subcol in col_names:
                        subcol_value = row[col_names[subcol] - 1].value if row[col_names[subcol] - 1].value else ""

                        if "_" in subcol:
                            # Split the column name to get the type suffix
                            subcol_type = subcol.split("_")[-1].strip().lower()
                            subcol = subcol[:-len(subcol_type)]

                            if subcol_type == "date":
                                sub_dict[subcol] = self._parse_date_(subcol_value)

                            elif subcol_type == "int":
                                try:
                                    sub_dict[subcol] = int(subcol_value)
                                except (ValueError, TypeError):
                                    raise InvalidTypeError(param, "int", subcol_value, worksheet_name, row[0].row)

                            elif subcol_type == "float":
                                try:
                                    sub_dict[subcol] = float(subcol_value)
                                except (ValueError, TypeError):
                                    raise InvalidTypeError(param, "float", subcol_value, worksheet_name, row[0].row)

                            elif subcol_type == "bool":
                                sub_dict[subcol] = True if subcol_value.lower() in ["true", "yes", "1"] else False

                            else:
                                raise InvalidTypeError(param, "valid type suffix", subcol_type, worksheet_name, row[0].row)

                        else: # treat as string - there is no type suffix
                            sub_dict[subcol] = subcol_value if subcol_value else ""
                    
                    # save list (complex) value
                    output_dict[param] = sub_dict

                ## Handle config reference recursion
                elif param_type == "config":
                    ref_sheet_name = value.strip().lower()
                    if ref_sheet_name != "":
                        output_dict[param] = self._parse_worksheet_(ref_sheet_name)
                    else:
                        raise ConfigReferenceError(param, worksheet_name, row[0].row)
                else:
                    raise InvalidTypeError(param, "valid type", param_type, worksheet_name, row[0].row)
            
        # finished - turn in home work
        return output_dict
            

        
        
            
    def _parse_date_(self, date_str: str):
        """Parse a date string into a date object."""
        if not date_str:
            return None
            
        # First check if it's already a datetime
        if isinstance(date_str, datetime.datetime):
            return date_str.date()
            
        # Try various date formats
        date_formats = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%m-%Y",
            "%d %b %Y",
            "%d %B %Y"
        ]
        
        for fmt in date_formats:
            try:
                return datetime.datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
                
        return None
    


    def _get_worksheet_(self, sheet_name: str):
        
        sheet_name = sheet_name.strip()
        
        """Retrieve a cached worksheet"""
        if sheet_name in self.worksheet.keys():
            return self.worksheet[sheet_name]

        """Check if a worksheet exists in the workbook."""
        if sheet_name is None or sheet_name.strip() == "":
            raise ValueError("Sheet name must be a non-empty string.")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Worksheet '{sheet_name}' does not exist in the workbook containing: ({', '.join(self.workbook.sheetnames)})")

        # cache the worksheet
        self.worksheet[sheet_name] = self.workbook[sheet_name]

        return self.worksheet[sheet_name] 




    # returns the worksheet, column index dict, and list of value column names
    def _get_worksheet_metaconfig_(self, sheet_name):
        
        # reset col_value_names
        col_value_names = []
        col_index = {}
        wksheet = self._get_worksheet_(sheet_name)
                
        # read first row to get column names
        first_row = [cell.value for cell in wksheet[1]]

        if ("param" not in first_row):
            raise MissingColumnError("param", sheet_name)
        if ("type" not in first_row):
            raise MissingColumnError("type", sheet_name)
        if ("value" not in first_row):
            raise MissingColumnError("value", sheet_name)
        
        # convert list to dict of column name to 1 based index
        col_index = {col_name: idx + 1 for idx, col_name in enumerate(first_row)} 
        print(col_index)
        # check the rows are the meta-config 
        if not wksheet.cell(row=2, column=col_index["type"]).value == "meta-config":
            raise ConfigXlsxReaderError(f"Second row of sheet '{sheet_name}' must be 'meta-config' row.")
        
        # get value columns
        col_value_names = [name.strip() for name in wksheet.cell(row=2, column=col_index["value"]).value.split(",")]

        return wksheet, col_index, col_value_names

        
        
        
        
        
        
    
    def _process_meta_sheet_(self, sheet_name: str):
        """Process the meta configuration sheet."""
        sheet = self.workbook[sheet_name]
        header_row = self._find_header_row_(sheet)
        
        if not header_row:
            return
        
        headers = self._get_row_values_(sheet, header_row)
        
        # Find columns for type, param, and value
        type_idx = headers.index('type') if 'type' in headers else None
        param_idx = headers.index('param') if 'param' in headers else None
        value_idx = headers.index('value') if 'value' in headers else None
        
        if not all([type_idx is not None, param_idx is not None, value_idx is not None]):
            return
            
        # Process the meta-config row (first data row after headers)
        meta_row = self._get_row_values_(sheet, header_row + 1)
        if len(meta_row) <= max(type_idx, param_idx, value_idx):
            return
            
        # Store meta-config for parsing other sheets
        self.meta_config = {
            "type_col": type_idx,
            "param_col": param_idx,
            "value_col": value_idx,
            # Other meta config as needed
        }
    
    


    def _find_header_row_(self, sheet):
        """Find the header row in the sheet."""
        for row_idx in range(1, min(10, sheet.max_row + 1)):  # Check first 10 rows
            row_values = self._get_row_values_(sheet, row_idx)
            if 'type' in row_values and 'param' in row_values:
                return row_idx
        return None
    
    def _get_row_values_(self, sheet, row_idx: int):
        """Get values from a row, converting to strings."""
        values = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[row_idx]]
        return values
    
    def _parse_value_columns_(self, value_str: str):
        """Parse value column string into column indices."""
        if not value_str or value_str.lower() == 'value':
            # Default to just the value column
            return []
            
        # Parse comma-separated column references
        columns = []
        for col_ref in value_str.split(','):
            col_ref = col_ref.strip()
            if col_ref:
                try:
                    # Handle both numeric indices and column letters (A, B, etc.)
                    if col_ref.isalpha():
                        columns.append(column_index_from_string(col_ref) - 1)  # 0-based index
                    else:
                        columns.append(int(col_ref) - 1)  # 0-based index
                except (ValueError, KeyError):
                    pass
        return columns
    

    
    def _get_cell_by_value_(self, sheet, value: str):
        """Find a cell with the given value in the sheet."""

        if not value:
            return None
            
        return next((cell for row in sheet.iter_rows() for cell in row if cell.value == value), None)
    

    ########################################
    
    def _extract_cell_style_(self, cell):
        """Extract style properties from a cell."""
        style = {}
        
        # Font properties
        if cell.font:
            style["font"] = {
                "name": cell.font.name,
                "size": cell.font.size,
                "bold": cell.font.bold,
                "italic": cell.font.italic,
                "color": cell.font.color.rgb if cell.font.color else None
            }
        
        # Fill properties
        if cell.fill:
            style["fill"] = {
                "type": cell.fill.fill_type,
                "start_color": cell.fill.start_color.rgb if cell.fill.start_color else None,
                "end_color": cell.fill.end_color.rgb if cell.fill.end_color else None
            }
        
        # Alignment properties
        if cell.alignment:
            style["alignment"] = {
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "wrap_text": cell.alignment.wrap_text,
                "text_rotation": cell.alignment.text_rotation
            }
        
        return style
    
    
    
    def _extract_border_style_(self, cell):
        """Extract border style properties from a cell."""
        border_style = {}
        
        if not cell.border:
            return border_style
            
        # Process each side of the border
        for side in ["top", "right", "bottom", "left"]:
            border_side = getattr(cell.border, side)
            if border_side:
                border_style[side] = {
                    "style": border_side.style,
                    "color": border_side.color.rgb if border_side.color else None
                }
                
        return border_style
